import argparse
import csv
from pathlib import Path
from tempfile import NamedTemporaryFile

import httpx
from openpyxl import load_workbook

DEFAULT_URL = (
    "https://digital.nhs.uk/binaries/content/assets/website-assets/"
    "developer/personal-demographics-service-fhir-api/all-pds-data-v3.xlsx"
)
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "pds.csv"
DATE_COLUMNS = {"DATE_OF_BIRTH", "DATE_OF_DEATH"}
NHS_NUMBER_COLUMN = "NHS_NUMBER"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--url", default=DEFAULT_URL)
    args = parser.parse_args()

    with httpx.Client(follow_redirects=True, timeout=60) as client:
        response = client.get(args.url)
        response.raise_for_status()

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(response.content)
        tmp.flush()
        workbook = load_workbook(tmp.name, read_only=True, data_only=True)

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    nhs_idx = header.index(NHS_NUMBER_COLUMN)

    with OUTPUT_PATH.open("w", newline="") as out:
        writer = csv.writer(out)
        writer.writerow(header)
        count = 0
        for row in rows:
            if not row or row[nhs_idx] is None:
                continue
            writer.writerow([_format(value, header[i]) for i, value in enumerate(row)])
            count += 1

    print(f"Wrote {count} rows to {OUTPUT_PATH}")  # noqa: T201


def _format(value: object, column: str) -> str:
    if value is None:
        return ""
    if column in DATE_COLUMNS and hasattr(value, "strftime"):
        return value.strftime("%Y%m%d")
    return str(value)


if __name__ == "__main__":
    main()
